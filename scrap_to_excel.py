import requests
from lxml import etree
import lxml.html
import csv
import openpyxl

def parse(url):
	name_of_csv = url.split('/')[-1].split('.')[0]
	try:
		api = requests.get(url)
	except:
		return
	tree = lxml.html.document_fromstring(api.text)
	words = tree.xpath('/html/body/div[4]/div[1]/div[1]/div[1]/div[2]/ul/li/a/text()')
	print(words)
	return words

def main():
	url = "https://www.allscrabblewords.com/{number}-letter-words/"
	i = 2
	wb = openpyxl.Workbook()
	wb.remove(wb['Sheet'])
	while i <= 12:
		wb.create_sheet(title='Page {number}'.format(number=i))
		sheet = wb['Page {number}'.format(number=i)]
		words = parse(url.format(number=i))
		for word in words:
			cell = sheet.cell(row = words.index(word)+1, column = 1)
			cell.value = word
		i += 1
	wb.save('scrapping_xl.xls')

if __name__ == '__main__':
	main()